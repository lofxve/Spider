import requests
import re
from bs4 import BeautifulSoup
from openpyxl import workbook  # 写入Excel表所用
import openpyxl
def get_movie_top250_name(soup):
    targets = soup.find_all('span', class_="title")  # 用BeautifulSoup找寻一个内容为一个列表
    targets_name = re.findall(r'.*?title">(.*?)<\/span', str(targets))  # 用正则表达式去掉标签
    for each in targets_name:  # 剔除targets_name当中的别名
        if '\xa0' in each:
            targets_name.remove(each)
    return targets_name


def get_movie_top250_workers(soup):
    targets = soup.find_all('p', class_="")
    targets_workers = []
    for each in targets:
        targets_workers.append(
            each.text.replace('<p class="">', '').replace('\n  ', '').replace('\xa0', '').replace('\n ', ''))
    return targets_workers


def get_movie_top250_star(soup):
    targets = soup.find_all('div', class_="star")
    targets_star = re.findall(r'<span class="rating_num" property="v:average">(.*?)<\/span>', str(targets))
    return targets_star


def get_movie_top250_quote(soup):
    targets = soup.find_all('p', class_="quote")
    targets_quote = re.findall(r'<span class="inq">(.*?)<\/span>', str(targets))
    return targets_quote


def save_to_excel(name, workers, star, quote):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "电影名称"
    ws['B1'] = "工作人员"
    ws['C1'] = "评分"
    ws['D1'] = "描述"
    for i in range(len(name)):
        result = [name[i], workers[i], star[i], quote[i]]
        ws.append(result)
    wb.save("豆瓣电影TOP250.xlsx")


def main():
    numbers = 1
    name = []
    workers = []
    star = []
    quote = []
    result = []
    while numbers:
        url = 'https://movie.douban.com/top250?start={}&filter='.format(numbers - 1)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}
        res = requests.get(url, headers=headers)
        soup = BeautifulSoup(res.text, "html.parser")
        name_1 = get_movie_top250_name(soup)
        workers_1 = get_movie_top250_workers(soup)
        star_1 = get_movie_top250_star(soup)
        quote_1 = get_movie_top250_quote(soup)
        for i in range(len(name_1)):
            name.append(name_1[i])
            workers.append(workers_1[i])
            star.append(star_1[i])
            try:
                quote.append(quote_1[i])
            except:
                quote.append("")
        numbers += 25
        if numbers > 250:
            break
        print(name, workers, star, quote)
        save_to_excel(name, workers, star, quote)
if __name__ == '__main__':
    main()
